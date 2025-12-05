from openpyxl import load_workbook, Workbook
from typing import Union, Optional, Dict
import pandas as pd
import numpy as np


def read_excel_table_to_df(excel_file_path, sheet_name, column_range=None, start_row=1, end_row=25000):
    """
    Read a table from a specific sheet of an Excel workbook.

    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the sheet containing the table.
    - column_range (str, optional): Range of columns for reading the table.
    - start_row (int, optional): Starting row for reading the table. Default is 1.
    - end_row (int, optional): Ending row for reading the table. Default is 2500.

    Returns:
    - pd.DataFrame: DataFrame containing the table data.
    """
    try:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0,
                           usecols=column_range, nrows=end_row, skiprows=range(1, start_row))
        # df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=column_range, nrows=end_row, engine="openpyxl")

        # Replace NaN by None
        df = df.replace(np.nan, None)

        return df
    except ParserError as e:
        raise ValueError(
            f"Error reading Excel file: {e}. Probable cause: improperly defined column_range, start_row and/or end_row. Please check the columns and rows in the sheet.") from e


def update_excel_from_dataframe(excel_file_path: str, sheet_name: str, columns_range: str,
                                df: pd.DataFrame, df_to_excel_cols_names_dict: Optional[Dict[str, str]] = None) -> None:
    """
    Open an Excel file, update specified columns in the specified sheet
    based on a pandas DataFrame, and save the changes.

    Parameters:
    - excel_file_path (str): The path to the Excel file.
    - sheet_name (str): The name of the sheet to update.
    - columns_range (str): The range of columns to update (e.g., 'A:O').
    - df (pd.DataFrame): The pandas DataFrame containing the data for update.
    - df_to_excel_cols_names_dict (Optional[Dict[str, str]]): A dictionary for renaming DataFrame columns
      before updating Excel. If None, the renaming is omitted.

    Example:
    update_excel_from_dataframe('example.xlsx', 'Sheet1', 'A:O', df, df_to_excel_cols_names_dict=None)
    """
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        # Check if the specified sheet exists
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return

        # Select the specified sheet
        sheet = workbook[sheet_name]

        # Read the first row into a dictionary mapping column names to column letters
        cols_name_dict = {col[0].value.strip(
        ): col[0].column_letter for col in sheet[columns_range]}

        # Clear old data in the specified columns from row 2 upwards
        for row in range(2, sheet.max_row + 1):
            for col in sheet[columns_range]:
                col_letter = col[0].column_letter
                sheet[col_letter + str(row)].value = None

        # Rename DataFrame columns if df_to_excel_cols_names_dict is provided
        if df_to_excel_cols_names_dict:
            df.rename(columns=df_to_excel_cols_names_dict, inplace=True)

        # Iterate over each column in the DataFrame
        for column in df.columns:
            # Iterate over each row in the DataFrame
            for index, row in df.iterrows():
                # Determine the corresponding Excel row (add 2 because Excel rows are 1-based)
                excel_row = index + 2

                # Get the column letter from the dictionary
                # Remove leading and trailing whitespaces
                col_letter = cols_name_dict.get(column.strip())

                # Check if the column exists in the Excel sheet
                if col_letter:
                    # Update the Excel sheet with the value from the DataFrame
                    sheet[col_letter + str(excel_row)] = row[column]

        # Save the changes back to the Excel file
        workbook.save(excel_file_path)
        print("Data updated and old data cleared successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")


def replace_sheet_content(excel_file_path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Loads an existing Excel file and replaces the content of a specified sheet with data from a DataFrame.

    Args:
        excel_file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to replace.
        df (pd.DataFrame): The DataFrame containing the data to write.

    Returns:
        None: The function does not return a value but saves changes to the Excel file.
    """
    try:
        # Load existing Excel file or create a new one if it doesn't exist
        try:
            wb = load_workbook(excel_file_path)
        except FileNotFoundError:
            wb = Workbook()

        # Check if the sheet exists, delete if it does
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Write DataFrame columns as header
        for col_num, value in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=value)

        # Write DataFrame rows
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Save changes
        wb.save(excel_file_path)
        print(f"Data replaced in sheet '{sheet_name}' in '{excel_file_path}'")
    except Exception as e:
        print(f"Error: {e}")
